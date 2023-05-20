from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

#新しいワークシートを追加
ws2 = wb.create_sheet("sheet2")
for x in range(1,11):
    for y in range(1,11):
        ws2.cell(row=x, column=y, value=x*y)


wb.save("sample.xlsx")